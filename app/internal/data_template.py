from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from typing import Dict, Any
import PyPDF2
import io


def create_data_template(template: UploadFile) -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    if not sheet:
        sheet = workbook.create_sheet()

    sheet.title = "Data Template"

    fields = load_pdf_fields(template)

    print(f"Found {len(fields)} fields in the PDF template.")

    for idx, (field_name, field_value) in enumerate(fields.items()):
        print(f"Processing field: {field_name} with value: {field_value}")
        cell = sheet.cell(row=1, column=idx + 1)
        cell.value = field_name
        if field_value:
            cell = sheet.cell(row=2, column=idx + 1)
            cell.value = field_value.get("/V", "")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{template.filename}.xlsx"'
        },
    )


def load_pdf_fields(template: UploadFile) -> Dict[str, Any]:
    """Extracts field names from a PDF template."""
    pdf_reader = PyPDF2.PdfReader(template.file)
    fields = pdf_reader.get_fields()
    if not fields:
        raise ValueError("No fields found in the PDF template.")
    print(fields)
    return fields
